import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal
from zipfile import BadZipFile
from typing import Iterable

from django.db import transaction
from rest_framework import serializers

from inventory.models import Inventory, InventoryBarcode


HEADER_ALIASES = {
    "product_name": {"product_name", "product", "medicine", "drug_name", "item_name", "name"},
    "strength": {"strength", "dosage", "dose"},
    "quantity_on_hand": {
        "quantity_on_hand",
        "quantity",
        "qty",
        "stock",
        "opening_stock",
        "current_stock",
    },
    "min_threshold": {
        "min_threshold",
        "threshold",
        "minimum_stock",
        "reorder_level",
        "min_stock",
    },
    "barcode": {
        "barcode",
        "bar_code",
        "ean",
        "ean_13",
        "gtin",
        "product_code",
        "scan_code",
    },
}

REQUIRED_FIELDS = ("product_name", "strength", "quantity_on_hand")


@dataclass(frozen=True)
class OpeningBalanceRow:
    row_number: int
    product_name: str
    strength: str
    quantity_on_hand: int
    min_threshold: int
    barcode: str = ""


def import_opening_balance(file_obj, extension: str) -> dict:
    rows = parse_opening_balance(file_obj, extension)
    return apply_opening_balance_rows(rows)


def apply_opening_balance_rows(rows: list[OpeningBalanceRow]) -> dict:
    created_count = 0
    updated_count = 0
    barcode_count = 0

    with transaction.atomic():
        for row in rows:
            item = (
                Inventory.objects.select_for_update()
                .filter(product_name=row.product_name, strength=row.strength)
                .first()
            )
            if item is None:
                item = Inventory.objects.create(
                    product_name=row.product_name,
                    strength=row.strength,
                    quantity_on_hand=row.quantity_on_hand,
                    min_threshold=row.min_threshold,
                )
                created_count += 1
            else:
                item.quantity_on_hand = row.quantity_on_hand
                item.min_threshold = row.min_threshold
                item.save(update_fields=["quantity_on_hand", "min_threshold", "updated_at"])
                updated_count += 1

            if row.barcode:
                barcode, created = InventoryBarcode.objects.get_or_create(
                    barcode=row.barcode,
                    defaults={
                        "inventory_item": item,
                        "is_primary": not InventoryBarcode.objects.filter(
                            inventory_item=item
                        ).exists(),
                    },
                )
                if barcode.inventory_item_id != item.id:
                    raise serializers.ValidationError(
                        {
                            "rows": [
                                {
                                    "row": row.row_number,
                                    "errors": {
                                        "barcode": [
                                            "Barcode is already assigned to another inventory item."
                                        ]
                                    },
                                }
                            ]
                        }
                    )
                if created:
                    barcode_count += 1

    return {
        "status": "completed",
        "total_rows": len(rows),
        "created_count": created_count,
        "updated_count": updated_count,
        "barcode_count": barcode_count,
    }


def validate_opening_balance_barcode_conflicts(rows: list[OpeningBalanceRow]) -> None:
    rows_by_barcode = {row.barcode: row for row in rows if row.barcode}
    if not rows_by_barcode:
        return

    existing_barcodes = (
        InventoryBarcode.objects.select_related("inventory_item")
        .filter(barcode__in=rows_by_barcode.keys())
    )
    errors = []
    for barcode in existing_barcodes:
        row = rows_by_barcode[barcode.barcode]
        item = barcode.inventory_item
        if (
            item.product_name.lower() != row.product_name.lower()
            or item.strength.lower() != row.strength.lower()
        ):
            errors.append(
                {
                    "row": row.row_number,
                    "errors": {
                        "barcode": [
                            "Barcode is already assigned to another inventory item."
                        ]
                    },
                }
            )

    if errors:
        raise serializers.ValidationError({"rows": errors})


def parse_opening_balance(file_obj, extension: str) -> list[OpeningBalanceRow]:
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    if extension == ".csv":
        raw_rows = _read_csv_rows(file_obj)
    elif extension == ".xlsx":
        raw_rows = _read_xlsx_rows(file_obj)
    else:
        raise serializers.ValidationError({"file": ["Unsupported opening balance file type."]})

    rows = _validate_rows(raw_rows)

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    return rows


def _read_csv_rows(file_obj) -> list[tuple[int, list]]:
    content = file_obj.read()
    if isinstance(content, str):
        text = content
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1252")

    reader = csv.reader(io.StringIO(text))
    return [(index, row) for index, row in enumerate(reader, start=1)]


def _read_xlsx_rows(file_obj) -> list[tuple[int, list]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise serializers.ValidationError(
            {"file": ["Excel import requires the openpyxl package."]}
        ) from exc

    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise serializers.ValidationError({"file": ["Invalid XLSX file."]}) from exc

    worksheet = workbook.active
    try:
        return [
            (index, list(row))
            for index, row in enumerate(worksheet.iter_rows(values_only=True), start=1)
        ]
    finally:
        workbook.close()


def _validate_rows(raw_rows: Iterable[tuple[int, list]]) -> list[OpeningBalanceRow]:
    raw_rows = list(raw_rows)
    header_info = _find_header(raw_rows)
    if header_info is None:
        raise serializers.ValidationError({"file": ["Opening balance file is empty."]})

    header_number, header_row = header_info
    columns = _map_headers(header_row)
    missing_fields = [field for field in REQUIRED_FIELDS if field not in columns]
    if missing_fields:
        raise serializers.ValidationError(
            {
                "headers": [
                    "Missing required column(s): " + ", ".join(missing_fields) + "."
                ]
            }
        )

    errors = []
    rows = []
    seen_keys = {}
    seen_barcodes = {}

    for row_number, raw_row in raw_rows:
        if row_number <= header_number or _is_blank_row(raw_row):
            continue

        try:
            row = _build_row(row_number, raw_row, columns)
        except serializers.ValidationError as exc:
            errors.append({"row": row_number, "errors": exc.detail})
            continue

        duplicate_key = (row.product_name.lower(), row.strength.lower())
        if duplicate_key in seen_keys:
            errors.append(
                {
                    "row": row_number,
                    "errors": {
                        "non_field_errors": [
                            f"Duplicate product/strength also found on row {seen_keys[duplicate_key]}."
                        ]
                    },
                }
            )
            continue

        seen_keys[duplicate_key] = row_number

        if row.barcode:
            if row.barcode in seen_barcodes:
                errors.append(
                    {
                        "row": row_number,
                        "errors": {
                            "barcode": [
                                f"Duplicate barcode also found on row {seen_barcodes[row.barcode]}."
                            ]
                        },
                    }
                )
                continue
            seen_barcodes[row.barcode] = row_number

        rows.append(row)

    if errors:
        raise serializers.ValidationError({"rows": errors})

    if not rows:
        raise serializers.ValidationError({"file": ["Opening balance file contains no data rows."]})

    return rows


def _find_header(raw_rows: list[tuple[int, list]]) -> tuple[int, list] | None:
    for row_number, row in raw_rows:
        if not _is_blank_row(row):
            return row_number, row
    return None


def _map_headers(header_row: list) -> dict[str, int]:
    columns = {}
    alias_to_field = {
        alias: field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }

    for index, value in enumerate(header_row):
        normalized = _normalize_header(value)
        field = alias_to_field.get(normalized)
        if field and field not in columns:
            columns[field] = index

    return columns


def _build_row(row_number: int, raw_row: list, columns: dict[str, int]) -> OpeningBalanceRow:
    product_name = _required_text(raw_row, columns["product_name"], "product_name")
    strength = _required_text(raw_row, columns["strength"], "strength")
    quantity_on_hand = _required_integer(
        raw_row, columns["quantity_on_hand"], "quantity_on_hand"
    )
    min_threshold = 0
    if "min_threshold" in columns:
        min_threshold = _optional_integer(raw_row, columns["min_threshold"], "min_threshold")
    barcode = ""
    if "barcode" in columns:
        barcode = _optional_text(raw_row, columns["barcode"])

    field_errors = {}
    if quantity_on_hand < 0:
        field_errors["quantity_on_hand"] = ["quantity_on_hand cannot be negative."]
    if min_threshold < 0:
        field_errors["min_threshold"] = ["min_threshold cannot be negative."]
    if field_errors:
        raise serializers.ValidationError(field_errors)

    return OpeningBalanceRow(
        row_number=row_number,
        product_name=product_name,
        strength=strength,
        quantity_on_hand=quantity_on_hand,
        min_threshold=min_threshold,
        barcode=barcode,
    )


def _required_text(raw_row: list, index: int, field_name: str) -> str:
    value = _value_at(raw_row, index)
    text = "" if value is None else str(value).strip()
    if not text:
        raise serializers.ValidationError({field_name: [f"{field_name} is required."]})
    return text


def _optional_text(raw_row: list, index: int) -> str:
    value = _value_at(raw_row, index)
    return "" if value is None else str(value).strip()


def _required_integer(raw_row: list, index: int, field_name: str) -> int:
    value = _value_at(raw_row, index)
    if value in (None, ""):
        raise serializers.ValidationError({field_name: [f"{field_name} is required."]})
    return _parse_integer(value, field_name)


def _optional_integer(raw_row: list, index: int, field_name: str) -> int:
    value = _value_at(raw_row, index)
    if value in (None, ""):
        return 0
    return _parse_integer(value, field_name)


def _parse_integer(value, field_name: str) -> int:
    if isinstance(value, bool):
        raise serializers.ValidationError({field_name: [f"{field_name} must be an integer."]})

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise serializers.ValidationError({field_name: [f"{field_name} must be an integer."]})

    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        raise serializers.ValidationError({field_name: [f"{field_name} must be an integer."]})

    text = str(value).strip()
    if re.fullmatch(r"-?\d+", text):
        return int(text)

    raise serializers.ValidationError({field_name: [f"{field_name} must be an integer."]})


def _value_at(raw_row: list, index: int):
    return raw_row[index] if index < len(raw_row) else None


def _is_blank_row(row: list) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _normalize_header(value) -> str:
    text = "" if value is None else str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")
